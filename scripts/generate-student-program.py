"""학생과 프로그램을 매칭해 student_program.csv를 생성한다.

실행 예:
    python scripts/generate-student-program.py
    python scripts/generate-student-program.py --min-programs 1 --max-programs 3 --seed 20260730
"""

from __future__ import annotations

import argparse
import csv
import random
import calendar
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_STUDENT_FILE = PROJECT_DIR / "data" / "chuncheon_students.xlsx"
DEFAULT_PROGRAM_FILE = PROJECT_DIR / "data" / "programs.csv"
DEFAULT_OUTPUT_FILE = PROJECT_DIR / "data" / "student_program.csv"

OUTPUT_COLUMNS = [
    "learning_history_id",
    "stu_id",
    "program_id",
    "enrollment_date",
    "start_date",
    "completion_date",
    "attendance_rate",
    "interest_score",
    "understanding_score",
    "difficulty_level",
]

SCHOOL_PREFIX = {
    "초등학교": "e",
    "중학교": "m",
    "고등학교": "h",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="학생별 프로그램 수강 이력 CSV 생성")
    parser.add_argument("--students", type=Path, default=DEFAULT_STUDENT_FILE)
    parser.add_argument("--programs", type=Path, default=DEFAULT_PROGRAM_FILE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_FILE)
    parser.add_argument("--min-programs", type=int, default=1)
    parser.add_argument("--max-programs", type=int, default=3)
    parser.add_argument("--seed", type=int, default=20260730)
    parser.add_argument(
        "--as-of-date",
        type=date.fromisoformat,
        default=date.today(),
        help="수강 상태 판단 기준일(YYYY-MM-DD, 기본값: 오늘)",
    )
    return parser.parse_args()


def read_students(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    students = [dict(zip(headers, row)) for row in rows]
    workbook.close()
    return students


def read_programs(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def student_grade_code(student: dict[str, object]) -> str:
    grade = str(student["grade"]).strip().lower()
    if grade[:1] in {"e", "m", "h"}:
        return grade

    school_level = str(student["school_level"]).strip()
    try:
        return f"{SCHOOL_PREFIX[school_level]}{int(float(grade))}"
    except (KeyError, ValueError) as error:
        raise ValueError(
            f"{student['stu_id']}: 학년 코드를 만들 수 없습니다 "
            f"(school_level={school_level!r}, grade={grade!r})"
        ) from error


def is_interest_match(interest: str, program: dict[str, str]) -> bool:
    flags = {name: program.get(name, "").strip() == "1" for name in (
        "is_sw",
        "is_ai",
        "is_bio",
        "is_sw_ai_related",
        "is_bio_related",
    )}
    if interest == "SW":
        return flags["is_sw"] or flags["is_sw_ai_related"]
    if interest == "AI":
        return flags["is_ai"] or flags["is_sw_ai_related"]
    if interest == "바이오":
        return flags["is_bio"] or flags["is_bio_related"]
    return False


def parse_date(value: str) -> date | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def enrollment_date(program: dict[str, str], start: date, rng: random.Random) -> date:
    apply_start = parse_date(program.get("apply_start", ""))
    apply_end = parse_date(program.get("apply_end", ""))

    if apply_start and apply_end:
        last_day = min(apply_end, start)
        if apply_start <= last_day:
            return apply_start + timedelta(days=rng.randint(0, (last_day - apply_start).days))
    if apply_end and apply_end <= start:
        return apply_end
    if apply_start and apply_start <= start:
        return apply_start
    return start - timedelta(days=rng.randint(7, 30))


WEEKDAY_NUMBER = {
    "월": calendar.MONDAY,
    "화": calendar.TUESDAY,
    "수": calendar.WEDNESDAY,
    "목": calendar.THURSDAY,
    "금": calendar.FRIDAY,
    "토": calendar.SATURDAY,
    "일": calendar.SUNDAY,
}


def course_progress_rate(
    program: dict[str, str],
    start: date,
    completion: date,
    as_of_date: date,
) -> float | str:
    """완료 과정은 100%, 수강 중 과정은 진행 차시 비율을 반환한다."""
    if as_of_date >= completion:
        return 100.0
    if as_of_date < start:
        return ""

    class_weekdays = {
        WEEKDAY_NUMBER[value]
        for value in program.get("weekdays", "").split("|")
        if value in WEEKDAY_NUMBER
    }
    if class_weekdays:
        total_sessions = sum(
            1
            for day_number in range((completion - start).days + 1)
            if (start + timedelta(days=day_number)).weekday() in class_weekdays
        )
        attended_sessions = sum(
            1
            for day_number in range((as_of_date - start).days + 1)
            if (start + timedelta(days=day_number)).weekday() in class_weekdays
        )
    else:
        # 요일 정보가 없는 프로그램은 전체 운영 일수 기준으로 계산한다.
        total_sessions = (completion - start).days + 1
        attended_sessions = (as_of_date - start).days + 1

    return round(min(attended_sessions / total_sessions * 100, 100.0), 1)


def eligible_programs(
    student: dict[str, object],
    programs: list[dict[str, str]],
    as_of_date: date,
) -> list[tuple[dict[str, str], date, date]]:
    grade_code = student_grade_code(student)
    interest = str(student["interest_category"]).strip()
    matches = []

    for program in programs:
        target_grades = {
            code.strip().lower()
            for code in program.get("target_grade_codes", "").split("|")
            if code.strip()
        }
        start = parse_date(program.get("start_date", ""))
        completion = parse_date(program.get("end_date", ""))
        if (
            grade_code in target_grades
            and is_interest_match(interest, program)
            and start is not None
            and completion is not None
            and completion >= start
            and start <= as_of_date
        ):
            matches.append((program, start, completion))
    return matches


def generate_history(
    students: list[dict[str, object]],
    programs: list[dict[str, str]],
    min_programs: int,
    max_programs: int,
    as_of_date: date,
    rng: random.Random,
) -> list[dict[str, object]]:
    history = []

    for student in students:
        candidates = eligible_programs(student, programs, as_of_date)
        if not candidates:
            print(f"경고: {student['stu_id']}에 맞는 프로그램이 없어 건너뜁니다.")
            continue

        count = min(rng.randint(min_programs, max_programs), len(candidates))
        ongoing = [
            candidate
            for candidate in candidates
            if candidate[1] <= as_of_date < candidate[2]
        ]
        selected = []
        if ongoing:
            selected.append(rng.choice(ongoing))
        remaining = [candidate for candidate in candidates if candidate not in selected]
        selected.extend(rng.sample(remaining, min(count - len(selected), len(remaining))))

        for program, start, completion in selected:
            history.append({
                "stu_id": student["stu_id"],
                "program_id": program["program_id"],
                "enrollment_date": enrollment_date(program, start, rng).isoformat(),
                "start_date": start.isoformat(),
                "completion_date": completion.isoformat(),
                "attendance_rate": course_progress_rate(
                    program,
                    start,
                    completion,
                    as_of_date,
                ),
                "interest_score": rng.randint(3, 5),
                "understanding_score": rng.randint(2, 5),
                "difficulty_level": rng.randint(1, 5),
            })

    history.sort(key=lambda row: (str(row["stu_id"]), str(row["start_date"]), str(row["program_id"])))
    for number, row in enumerate(history, start=1):
        row["learning_history_id"] = f"LH_{number:04d}"
    return history


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    args = parse_args()
    if args.min_programs < 1 or args.max_programs < args.min_programs:
        raise ValueError("수강 개수는 1 이상이고 최댓값은 최솟값 이상이어야 합니다.")

    rng = random.Random(args.seed)
    students = read_students(args.students)
    programs = read_programs(args.programs)
    history = generate_history(
        students,
        programs,
        args.min_programs,
        args.max_programs,
        args.as_of_date,
        rng,
    )
    write_csv(args.output, history)
    print(f"{len(history)}건 생성: {args.output}")


if __name__ == "__main__":
    main()
